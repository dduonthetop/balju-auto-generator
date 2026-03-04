from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from pathlib import Path
import subprocess
from typing import Any, Dict, List, Optional, Tuple
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


SENDER_NAME = "스미후루코리아"
COURIER_NAME = "롯데택배"
BACKUP_KEEP_COUNT = 3


def is_temp_excel(path: Path) -> bool:
    return path.name.startswith("~$")


def find_target_file(base: Path) -> Path:
    candidates = [
        p
        for p in base.glob("*.xlsx")
        if not is_temp_excel(p)
        and not p.name.startswith("주식회사지엠_")
        and ".bak_" not in p.name
    ]
    if len(candidates) != 1:
        raise RuntimeError(f"타깃 파일은 1개여야 합니다. 현재: {len(candidates)}개")
    return candidates[0]


def find_target_by_name(base: Path, target_name: str) -> Path:
    target = base / target_name
    if not target.exists():
        raise RuntimeError(f"지정한 타깃 파일이 없습니다: {target_name}")
    return target


def find_source_files(base: Path, source_date: Optional[str] = None) -> List[Path]:
    files = sorted(
        [
            p
            for p in base.glob("*.xlsx")
            if not is_temp_excel(p) and p.name.startswith("주식회사지엠_")
        ]
    )
    if source_date:
        files = [p for p in files if source_date in p.name]
    return files


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="연테이블 공동구매 기준 파일에 주식회사지엠 주문 데이터를 자동 반영합니다."
    )
    parser.add_argument(
        "--target",
        help="기준 파일명 (예: 연테이블 공동구매 3월4일.xlsx). 미지정 시 폴더 내 자동 탐색(1개여야 함).",
    )
    parser.add_argument(
        "--source-date",
        help="소스 파일명 날짜 필터(예: 20260304). 지정 시 파일명에 해당 문자열이 포함된 소스만 사용.",
    )
    parser.add_argument(
        "--backup-keep",
        type=int,
        default=BACKUP_KEEP_COUNT,
        help=f"유지할 최신 백업 파일 개수(기본 {BACKUP_KEEP_COUNT})",
    )
    parser.add_argument(
        "--base-dir",
        help="작업 폴더 경로. 미지정 시 apply_mapping.py 파일이 있는 폴더를 사용.",
    )
    return parser.parse_args()


def cleanup_old_backups(target_file: Path, keep_count: int) -> int:
    pattern = f"{target_file.stem}.bak_*{target_file.suffix}"
    backups = sorted(target_file.parent.glob(pattern), key=lambda p: p.name)
    if keep_count < 0:
        keep_count = 0
    to_delete = backups[:-keep_count] if keep_count > 0 else backups
    deleted = 0
    for p in to_delete:
        try:
            p.unlink()
            deleted += 1
        except Exception:
            pass
    return deleted


def map_option_to_spec(option_text: str) -> Optional[str]:
    text = (option_text or "").strip()
    if not text:
        return None

    if "3종 혼합" in text:
        return "바나밥 시즈닝 바나나칩 3종(어니언1+솔티드1+김1)-03ea"
    if "김맛" in text:
        return "바나밥 시즈닝 바나나칩 김맛-09ea"
    if "샤워크림" in text or "어니언" in text:
        return "바나밥 시즈닝 바나나칩 샤워크림앤어니언맛-09ea"
    if "솔티드카라멜" in text:
        return "바나밥 시즈닝 바나나칩 솔티드카라멜맛-09ea"
    if "쿠키슈" in text:
        return "쿠키슈 4입-02ea"
    if "바삭 바나나칩" in text or "바나나칩 70g x 8입" in text:
        return "바나나칩-08ea"
    return None


def read_price_table(target_wb) -> Dict[str, int]:
    ws = target_wb[target_wb.sheetnames[1]]
    price_table: Dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        spec_name = ws.cell(row=r, column=1).value
        supply_price = ws.cell(row=r, column=7).value
        if spec_name in (None, "", "TTL"):
            continue
        if isinstance(spec_name, str):
            spec_name = spec_name.strip()
        if spec_name and isinstance(supply_price, (int, float)):
            price_table[str(spec_name)] = int(supply_price)
    return price_table


def extract_source_rows(source_files: List[Path]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for source_path in source_files:
        wb = load_workbook(source_path, data_only=True)
        if "Supply" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Supply"]
        for r in range(2, ws.max_row + 1):
            order_no = ws.cell(row=r, column=1).value
            if order_no in (None, ""):
                continue
            row = {
                "주문번호": ws.cell(row=r, column=1).value,
                "주문상품고유번호": ws.cell(row=r, column=3).value,
                "수령인명": ws.cell(row=r, column=6).value,
                "수령인연락처": ws.cell(row=r, column=7).value,
                "우편번호": ws.cell(row=r, column=8).value,
                "주소": ws.cell(row=r, column=9).value,
                "상품코드": ws.cell(row=r, column=10).value,
                "고객선택옵션": ws.cell(row=r, column=12).value,
                "주문수량": ws.cell(row=r, column=13).value,
                "배송시 요청사항": ws.cell(row=r, column=16).value,
            }
            rows.append(row)
        wb.close()

    dedup: Dict[Any, Dict[str, Any]] = {}
    for row in rows:
        key = row["주문상품고유번호"]
        dedup[key] = row

    merged_rows = list(dedup.values())
    merged_rows.sort(
        key=lambda x: (
            str(x.get("주문번호") or ""),
            str(x.get("주문상품고유번호") or ""),
        )
    )
    return merged_rows


def build_output_rows(
    source_rows: List[Dict[str, Any]],
    price_table: Dict[str, int],
) -> Tuple[List[List[Any]], List[Tuple[Optional[str], int, Optional[int]]], List[str]]:
    output_rows: List[List[Any]] = []
    detail_rows_for_summary: List[Tuple[Optional[str], int, Optional[int]]] = []
    warnings: List[str] = []

    for row in source_rows:
        option_name = str(row.get("고객선택옵션") or "").strip()
        qty_raw = row.get("주문수량")
        try:
            qty = int(qty_raw) if qty_raw is not None else 0
        except Exception:
            qty = 0

        spec_name = map_option_to_spec(option_name)
        supply_price: Optional[int] = price_table.get(spec_name) if spec_name else None
        item_name = spec_name if spec_name else option_name
        if qty <= 0:
            warnings.append(
                f"[수량이상] 주문상품고유번호={row.get('주문상품고유번호')} 주문수량={qty_raw}"
            )
            continue

        amount_per_row = supply_price if supply_price is not None else None

        if spec_name is None:
            warnings.append(
                f"[품명매핑실패] 주문상품고유번호={row.get('주문상품고유번호')} 옵션='{option_name}'"
            )
        elif supply_price is None:
            warnings.append(
                f"[공급가없음] 주문상품고유번호={row.get('주문상품고유번호')} 스펙='{spec_name}'"
            )

        # 기준 파일 규칙: 모든 상세 라인은 수량 1로 분해해서 기록
        for _ in range(qty):
            detail_rows_for_summary.append((spec_name, 1, amount_per_row))
            output_rows.append(
                [
                    SENDER_NAME,  # A 보내는분
                    row.get("수령인명"),  # B 받는사람
                    row.get("수령인연락처"),  # C 받으시는분 전화
                    row.get("주소"),  # D 받는분 총주소
                    1,  # E 수량(고정)
                    item_name,  # F 품목명(품명 및 규격 기준)
                    str(row.get("우편번호") or ""),  # G 우편번호
                    row.get("배송시 요청사항"),  # H 특이사항
                    amount_per_row,  # I 금액(1개 단가)
                    None,  # J 송장번호 (공란)
                    row.get("주문번호"),  # K 주문번호
                    row.get("주문상품고유번호"),  # L 주문상품고유번호
                    row.get("상품코드"),  # M 상품코드
                    COURIER_NAME,  # N 택배사
                    None,  # O 배송번호 (공란)
                ]
            )

    return output_rows, detail_rows_for_summary, warnings


def write_rows_to_target(ws, output_rows: List[List[Any]]) -> None:
    start_row = 2
    for idx, row_values in enumerate(output_rows):
        r = start_row + idx
        for c, value in enumerate(row_values, start=1):
            ws.cell(row=r, column=c).value = value


def apply_template_style_to_rows(ws, template_row: int, start_row: int, row_count: int) -> None:
    template_styles = {}
    for c in range(1, 16):
        template_cell = ws.cell(row=template_row, column=c)
        template_styles[c] = copy(template_cell._style)

    template_height = ws.row_dimensions[template_row].height
    for offset in range(row_count):
        r = start_row + offset
        for c in range(1, 16):
            ws.cell(row=r, column=c)._style = copy(template_styles[c])
        ws.row_dimensions[r].height = template_height


def apply_uniform_style_from_anchor(
    ws,
    anchor_row: int,
    anchor_col: int,
    start_row: int,
    row_count: int,
) -> None:
    anchor_style = copy(ws.cell(row=anchor_row, column=anchor_col)._style)
    anchor_height = ws.row_dimensions[anchor_row].height
    for offset in range(row_count):
        r = start_row + offset
        for c in range(1, 16):
            ws.cell(row=r, column=c)._style = copy(anchor_style)
        ws.row_dimensions[r].height = anchor_height


def find_style_template_row(ws, min_row: int = 2, max_row: int = 200) -> int:
    upper = min(max_row, ws.max_row)
    for r in range(min_row, upper + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        if any(v not in (None, "") for v in values):
            continue
        style_ids = [ws.cell(row=r, column=c).style_id for c in range(1, 16)]
        if all(style_ids[c - 1] != 0 for c in range(2, 11)):
            return r
    return 20 if ws.max_row >= 20 else 2


def get_last_used_row(ws) -> int:
    last = 1
    for (r, c), cell in ws._cells.items():
        if 1 <= c <= 15 and cell.value not in (None, ""):
            last = max(last, r)
    return last


def clear_rows_values_only(ws, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    for r in range(start_row, end_row + 1):
        for c in range(1, 16):
            ws.cell(row=r, column=c).value = None


def apply_display_format(ws, data_row_count: int) -> None:
    no_fill = PatternFill(fill_type=None)

    # 1행 제외(A:O) 배경색 없음
    for r in range(2, ws.max_row + 1):
        for c in range(1, 16):
            ws.cell(row=r, column=c).fill = copy(no_fill)

    # 금액(I열) KRW 표시 형식
    for r in range(2, 2 + data_row_count):
        ws.cell(row=r, column=9).number_format = "₩#,##0"


def update_summary_sheet(
    ws_summary,
    detail_rows_for_summary: List[Tuple[Optional[str], int, Optional[int]]],
) -> None:
    # 헤더 고정 복구
    ws_summary.cell(row=2, column=8).value = "발주수량"
    ws_summary.cell(row=2, column=9).value = "금액(원)"

    spec_row_map: Dict[str, int] = {}
    ttl_row: Optional[int] = None

    for r in range(1, ws_summary.max_row + 1):
        name_val = ws_summary.cell(row=r, column=1).value
        if not isinstance(name_val, str):
            continue
        key = name_val.strip()
        if key == "TTL":
            ttl_row = r
            continue
        if r <= 2:
            continue
        supply_val = ws_summary.cell(row=r, column=7).value
        if not isinstance(supply_val, (int, float)):
            continue
        if key:
            spec_row_map[key] = r

    qty_sum_by_spec: Dict[str, int] = {}
    amt_sum_by_spec: Dict[str, int] = {}
    for spec_name, qty, amount in detail_rows_for_summary:
        if not spec_name:
            continue
        qty_sum_by_spec[spec_name] = qty_sum_by_spec.get(spec_name, 0) + qty
        amt_sum_by_spec[spec_name] = amt_sum_by_spec.get(spec_name, 0) + int(amount or 0)

    for spec_name, row_idx in spec_row_map.items():
        q = qty_sum_by_spec.get(spec_name, 0)
        a = amt_sum_by_spec.get(spec_name, 0)
        ws_summary.cell(row=row_idx, column=8).value = q
        ws_summary.cell(row=row_idx, column=9).value = a
        ws_summary.cell(row=row_idx, column=9).number_format = "₩#,##0"

    if ttl_row is not None:
        data_rows = sorted(spec_row_map.values())
        if data_rows:
            first_row = data_rows[0]
            last_row = data_rows[-1]
            ws_summary.cell(row=ttl_row, column=8).value = f"=SUM(H{first_row}:H{last_row})"
            ws_summary.cell(row=ttl_row, column=9).value = f"=SUM(I{first_row}:I{last_row})"
            ws_summary.cell(row=ttl_row, column=9).number_format = "₩#,##0"


def autofit_columns(
    ws,
    min_col: int,
    max_col: int,
    max_row: int,
    min_width: float = 8.0,
) -> None:
    def visual_width(text: str) -> int:
        width = 0
        for ch in text:
            ea = unicodedata.east_asian_width(ch)
            width += 2 if ea in ("F", "W") else 1
        return width

    for c in range(min_col, max_col + 1):
        max_len = 0.0
        for r in range(1, max_row + 1):
            value = ws.cell(row=r, column=c).value
            if value is None:
                continue
            text = str(value).replace("\n", " ")
            length = float(visual_width(text))
            if length > max_len:
                max_len = length

        # 상한 없이 문자열 표시 폭 기준으로 너비 계산
        width = max(min_width, (max_len * 1.15) + 2)
        ws.column_dimensions[get_column_letter(c)].width = width


def excel_autofit_columns_via_com(target_file: Path) -> bool:
    escaped = str(target_file.resolve()).replace("'", "''")
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
try {{
    $wb = $excel.Workbooks.Open('{escaped}')
    foreach ($ws in $wb.Worksheets) {{
        $null = $ws.UsedRange
        $ws.UsedRange.Columns.AutoFit() | Out-Null
    }}
    $wb.Save()
    $wb.Close($true)
}} finally {{
    if ($wb) {{ [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) }}
    $excel.Quit()
    [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel)
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_script],
            check=True,
            capture_output=True,
            text=True,
        )
        return True
    except Exception:
        return False


def main() -> None:
    args = parse_args()
    base = Path(args.base_dir).resolve() if args.base_dir else Path(__file__).resolve().parent
    target_file = (
        find_target_by_name(base, args.target) if args.target else find_target_file(base)
    )
    source_files = find_source_files(base, source_date=args.source_date)
    if not source_files:
        if args.source_date:
            raise RuntimeError(
                f"소스 파일을 찾지 못했습니다. 패턴=주식회사지엠_*.xlsx, 날짜필터={args.source_date}"
            )
        raise RuntimeError("소스 파일(주식회사지엠_*.xlsx)을 찾지 못했습니다.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = target_file.with_name(f"{target_file.stem}.bak_{timestamp}{target_file.suffix}")
    backup_path.write_bytes(target_file.read_bytes())

    wb_target = load_workbook(target_file)
    ws_detail = wb_target[wb_target.sheetnames[0]]
    ws_summary = wb_target[wb_target.sheetnames[1]]

    price_table = read_price_table(wb_target)
    source_rows = extract_source_rows(source_files)
    output_rows, detail_rows_for_summary, warnings = build_output_rows(source_rows, price_table)

    old_last_row = get_last_used_row(ws_detail)
    template_row = find_style_template_row(ws_detail)
    write_rows_to_target(ws_detail, output_rows)
    apply_uniform_style_from_anchor(
        ws_detail,
        anchor_row=2,
        anchor_col=1,
        start_row=2,
        row_count=len(output_rows),
    )
    new_last_row = 1 + len(output_rows)
    clear_rows_values_only(ws_detail, start_row=new_last_row + 1, end_row=old_last_row)
    apply_display_format(ws_detail, data_row_count=len(output_rows))
    update_summary_sheet(ws_summary, detail_rows_for_summary)
    autofit_columns(ws_detail, min_col=1, max_col=15, max_row=max(ws_detail.max_row, new_last_row))
    autofit_columns(ws_summary, min_col=1, max_col=9, max_row=ws_summary.max_row)

    wb_target.save(target_file)
    wb_target.close()
    deleted_backups = cleanup_old_backups(target_file, args.backup_keep)
    com_autofit_ok = excel_autofit_columns_via_com(target_file)

    print(f"타깃 파일: {target_file.name}")
    print(f"백업 파일: {backup_path.name}")
    print(f"백업 정리: {deleted_backups}개 삭제, 최신 {args.backup_keep}개 유지")
    print(f"소스 파일 수: {len(source_files)}")
    if args.source_date:
        print(f"소스 날짜 필터: {args.source_date}")
    print(f"반영 행 수: {len(output_rows)}")
    print(f"엑셀 AutoFit(더블클릭 동일): {'성공' if com_autofit_ok else '실패(내장 계산식 유지)'}")
    if warnings:
        print(f"경고 수: {len(warnings)}")
        for item in warnings:
            print(item)
    else:
        print("경고 수: 0")


if __name__ == "__main__":
    main()
